"""Simplified spreadsheet importer for core operational data."""
from __future__ import annotations

import io
import re
from collections import defaultdict
from dataclasses import dataclass, field
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation
from typing import Any, Iterable, Mapping

from fastapi import HTTPException, status

try:  # pragma: no cover - optional dependency
    from openpyxl import load_workbook
except Exception:  # pragma: no cover - optional dependency
    load_workbook = None

from sqlalchemy import delete, select
from sqlalchemy.ext.asyncio import AsyncSession

from ..db import engine
from ..models import domain
from ..models.base import Base
from ..utils.datetime import utc_now

SUPPORTED_SHEETS = {"products", "customers", "orders", "purchase_orders", "vendors"}

NO_IMPORTABLE_ROWS_WARNING = "No importable rows were found in the spreadsheet."

FIELD_ALIASES: dict[str, dict[str, set[str]]] = {
    "products": {
        "sku": {"sku", "product_sku", "item_sku", "item_number", "item_id"},
        "description": {"description", "product_description", "name", "product_name", "item_name", "item_description"},
        "category": {"category"},
        "subcategory": {"subcategory", "sub_category"},
        "unit_cost": {"unit_cost", "cost"},
        "price": {"price", "retail", "sale_price"},
        "tax_code": {"tax_code"},
        "barcode": {"barcode", "upc"},
        "qty_on_hand": {"qty_on_hand", "quantity", "qty", "on_hand", "inventory"},
        "location_name": {"location", "location_name", "warehouse", "store", "site"},
        "vendor_name": {"vendor", "vendor_name"},
    },
    "customers": {
        "name": {"name", "customer_name"},
        "email": {"email", "customer_email"},
        "phone": {"phone", "customer_phone"},
    },
    "orders": {
        "external_ref": {"order_number", "order_no", "order_id", "external_ref"},
        "status": {"status"},
        "subtotal": {"subtotal"},
        "tax": {"tax"},
        "total": {"total", "order_total"},
        "deposit_amt": {"deposit", "deposit_amt"},
        "order_date": {"order_date", "sale_date", "date"},
        "created_at": {"created_at", "created_on"},
        "created_by": {"created_by", "sales_rep", "owner"},
        "notes": {"notes", "order_notes"},
        "customer_email": {"customer_email", "email"},
        "customer_name": {"customer_name", "name"},
        "customer_phone": {"customer_phone", "phone"},
        "item_sku": {"item_sku", "product_sku", "sku"},
        "qty": {"qty", "quantity"},
        "unit_price": {"unit_price", "line_price"},
        "location_name": {"location", "location_name"},
    },
    "purchase_orders": {
        "external_ref": {"po_number", "po_no", "reference", "external_ref"},
        "vendor_name": {"vendor_name", "vendor"},
        "status": {"status"},
        "expected_date": {"expected_date", "eta", "due_date"},
        "created_at": {"created_at", "created_on"},
        "created_by": {"created_by", "buyer"},
        "terms": {"terms"},
        "notes": {"notes"},
        "item_sku": {"item_sku", "product_sku", "sku"},
        "item_description": {"item_description", "description"},
        "qty_ordered": {"qty_ordered", "quantity", "qty"},
        "unit_cost": {"unit_cost", "cost"},
    },
    "vendors": {
        "name": {"vendor_name", "name"},
        "email": {"vendor_email", "email"},
        "phone": {"vendor_phone", "phone"},
        "terms": {"terms"},
        "address_line1": {"address", "address_line1", "street"},
        "address_line2": {"address_line2", "suite", "apt"},
        "city": {"city"},
        "state": {"state", "province", "region"},
        "postal_code": {"postal_code", "zip", "zip_code"},
        "country": {"country"},
    },
}


@dataclass
class ImportCounters:
    vendors: int = 0
    locations: int = 0
    items: int = 0
    barcodes: int = 0
    inventory_records: int = 0
    customers: int = 0
    sales: int = 0
    purchase_orders: int = 0
    receivings: int = 0
    warnings: list[str] = field(default_factory=list)


@dataclass
class ImportResult:
    counters: ImportCounters
    cleared_sample_data: bool
    imported_at: datetime


async def import_spreadsheet(
    session: AsyncSession, data: bytes, filename: str, dataset: str | None = None
) -> ImportResult:
    dataset_key = dataset.lower() if dataset else None
    if dataset_key is not None and dataset_key not in SUPPORTED_SHEETS:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=f"Unsupported dataset '{dataset}'.",
        )

    datasets = extract_datasets(data, filename, preferred_entity=dataset_key)
    counters = ImportCounters()

    if dataset_key is not None:
        if not datasets.get(dataset_key):
            counters.warnings.append(NO_IMPORTABLE_ROWS_WARNING)
            return ImportResult(
                counters=counters,
                cleared_sample_data=False,
                imported_at=utc_now(),
            )
    elif not any(datasets.get(name) for name in SUPPORTED_SHEETS):
        counters.warnings.append(NO_IMPORTABLE_ROWS_WARNING)
        return ImportResult(
            counters=counters,
            cleared_sample_data=False,
            imported_at=utc_now(),
        )

    should_clear_demo = dataset_key in (None, "products")
    cleared_demo = False
    vendor_index: dict[str, domain.Vendor]
    location_index: dict[str, domain.Location]
    existing_items: dict[str, domain.Item]
    short_codes_in_use: set[str]
    customers_index: dict[str, domain.Customer]

    if should_clear_demo:
        cleared_demo = await _clear_existing_data(session)
        vendor_index = {}
        location_index = {}
        existing_items = {}
        short_codes_in_use = set()
        customers_index = {}
    else:
        await _ensure_schema()
        vendor_index = await _load_vendor_index(session)
        location_index = await _load_location_index(session)
        existing_items, short_codes_in_use = await _load_item_index(session)
        customers_index = await _load_customer_index(session)

    vendor_rows: Iterable[Mapping[str, Any]]
    if dataset_key in (None, "vendors"):
        vendor_rows = datasets.get("vendors", [])
    else:
        vendor_rows = []

    if vendor_rows:
        vendor_index = await _import_vendors(session, vendor_rows, counters, vendor_index)

    product_rows: Iterable[Mapping[str, Any]]
    if dataset_key in (None, "products"):
        product_rows = datasets.get("products", [])
    else:
        product_rows = []

    items_index = existing_items
    if product_rows:
        items_index = await _import_products(
            session,
            product_rows,
            counters,
            vendor_index,
            location_index,
            existing_items=existing_items if not should_clear_demo else None,
            short_codes=short_codes_in_use if not should_clear_demo else None,
        )

    customer_rows: Iterable[Mapping[str, Any]]
    if dataset_key in (None, "customers"):
        customer_rows = datasets.get("customers", [])
    else:
        customer_rows = []

    if customer_rows:
        customers_index = await _import_customers(
            session,
            customer_rows,
            counters,
            existing=customers_index if not should_clear_demo else None,
        )

    if dataset_key in (None, "orders"):
        await _import_orders(
            session,
            datasets.get("orders", []),
            counters,
            customers_index,
            items_index,
            location_index,
        )

    if dataset_key in (None, "purchase_orders"):
        await _import_purchase_orders(
            session,
            datasets.get("purchase_orders", []),
            counters,
            vendor_index,
            items_index,
        )

    await session.flush()

    return ImportResult(
        counters=counters,
        cleared_sample_data=cleared_demo,
        imported_at=utc_now(),
    )


async def _import_vendors(
    session: AsyncSession,
    rows: Iterable[Mapping[str, Any]],
    counters: ImportCounters,
    vendor_index: dict[str, domain.Vendor],
) -> dict[str, domain.Vendor]:
    for row in rows:
        name = _coerce_str(row.get("name"))
        if not name:
            counters.warnings.append("Skipped vendor row without a name")
            continue

        key = name.lower()
        email = _coerce_str(row.get("email"))
        phone = _coerce_str(row.get("phone"))
        terms = _coerce_str(row.get("terms"))
        address_components = {
            "line1": _coerce_str(row.get("address_line1")),
            "line2": _coerce_str(row.get("address_line2")),
            "city": _coerce_str(row.get("city")),
            "state": _coerce_str(row.get("state")),
            "postal_code": _coerce_str(row.get("postal_code")),
            "country": _coerce_str(row.get("country")),
        }
        address_json = {k: v for k, v in address_components.items() if v}

        vendor = vendor_index.get(key)
        if vendor is None:
            vendor = domain.Vendor(
                name=name,
                terms=terms,
                phone=phone,
                email=email,
                address_json=address_json or None,
                active=True,
            )
            session.add(vendor)
            await session.flush()
            vendor_index[key] = vendor
            counters.vendors += 1
        else:
            updated = False
            if email and vendor.email != email:
                vendor.email = email
                updated = True
            if phone and vendor.phone != phone:
                vendor.phone = phone
                updated = True
            if terms and vendor.terms != terms:
                vendor.terms = terms
                updated = True
            if address_json and vendor.address_json != address_json:
                vendor.address_json = address_json
                updated = True
            if updated:
                session.add(vendor)

    return vendor_index


async def _import_products(
    session: AsyncSession,
    rows: Iterable[Mapping[str, Any]],
    counters: ImportCounters,
    vendor_index: dict[str, domain.Vendor],
    location_index: dict[str, domain.Location],
    *,
    existing_items: dict[str, domain.Item] | None = None,
    short_codes: set[str] | None = None,
) -> dict[str, domain.Item]:
    if existing_items is None:
        items: dict[str, domain.Item] = {}
    else:
        items = dict(existing_items)
    short_codes_in_use: set[str]
    if short_codes is not None:
        short_codes_in_use = set(short_codes)
    else:
        short_codes_in_use = {item.short_code for item in items.values()}

    for row in rows:
        sku = _coerce_str(row.get("sku"))
        description = _coerce_str(row.get("description"))
        if not sku or not description:
            counters.warnings.append("Skipped product row without SKU and description")
            continue

        sku_key = sku.lower()
        existing_item = items.get(sku_key)
        is_new_item = existing_item is None

        unit_cost = _coerce_decimal(row.get("unit_cost"))
        price = _coerce_decimal(row.get("price"))
        if price is None:
            price = unit_cost or Decimal("0")
        if unit_cost is None:
            unit_cost = price or Decimal("0")

        category = _coerce_str(row.get("category"))
        subcategory = _coerce_str(row.get("subcategory"))
        tax_code = _coerce_str(row.get("tax_code"))

        if existing_item is not None:
            if existing_items is not None and sku_key in existing_items:
                updated = False
                if existing_item.description != description:
                    existing_item.description = description
                    updated = True
                if category and existing_item.category != category:
                    existing_item.category = category
                    updated = True
                if subcategory and existing_item.subcategory != subcategory:
                    existing_item.subcategory = subcategory
                    updated = True
                if unit_cost is not None and existing_item.unit_cost != unit_cost:
                    existing_item.unit_cost = unit_cost
                    updated = True
                if price is not None and existing_item.price != price:
                    existing_item.price = price
                    updated = True
                if tax_code and existing_item.tax_code != tax_code:
                    existing_item.tax_code = tax_code
                    updated = True
                if updated:
                    session.add(existing_item)
                item = existing_item
            else:
                counters.warnings.append(f"Duplicate product with SKU '{sku}' skipped")
                continue
        else:
            short_code = _generate_short_code(sku, short_codes_in_use)
            item = domain.Item(
                sku=sku,
                description=description,
                category=category,
                subcategory=subcategory,
                unit_cost=unit_cost,
                price=price,
                tax_code=tax_code,
                short_code=short_code,
                active=True,
            )
            session.add(item)
            await session.flush()
            items[sku_key] = item
            counters.items += 1

        if is_new_item:
            short_codes_in_use.add(item.short_code)

        vendor_name = _coerce_str(row.get("vendor_name"))
        if vendor_name:
            await _get_or_create_vendor(session, vendor_name, vendor_index, counters)

        barcode_value = _coerce_str(row.get("barcode"))
        if barcode_value:
            session.add(domain.Barcode(item_id=item.item_id, barcode=barcode_value))
            counters.barcodes += 1

        qty = _coerce_decimal(row.get("qty_on_hand"))
        if qty is not None and qty != Decimal("0"):
            location_name = _coerce_str(row.get("location_name")) or "Main Warehouse"
            location = await _get_or_create_location(
                session, location_name, location_index, counters
            )
            inventory = domain.Inventory(
                item_id=item.item_id,
                location_id=location.location_id,
                qty_on_hand=qty,
                qty_reserved=Decimal("0"),
                avg_cost=unit_cost,
            )
            session.add(inventory)
            counters.inventory_records += 1

    return items


async def _import_customers(
    session: AsyncSession,
    rows: Iterable[Mapping[str, Any]],
    counters: ImportCounters,
    *,
    existing: dict[str, domain.Customer] | None = None,
) -> dict[str, domain.Customer]:
    customers: dict[str, domain.Customer] = dict(existing or {})

    for row in rows:
        name = _coerce_str(row.get("name"))
        email = _coerce_str(row.get("email"))
        phone = _coerce_str(row.get("phone"))
        if not any([name, email, phone]):
            counters.warnings.append("Skipped customer row without identifying fields")
            continue

        key_candidates = _customer_lookup_keys(name, email, phone)
        existing_customer = None
        for candidate in key_candidates:
            existing_customer = customers.get(candidate)
            if existing_customer is not None:
                break

        if existing_customer is not None:
            updated = False
            if name and existing_customer.name != name:
                existing_customer.name = name
                updated = True
            if email and existing_customer.email != email:
                existing_customer.email = email
                updated = True
            if phone and existing_customer.phone != phone:
                existing_customer.phone = phone
                updated = True
            if updated:
                session.add(existing_customer)
            for candidate in key_candidates:
                if candidate:
                    customers[candidate] = existing_customer
            continue

        customer = domain.Customer(
            name=name or email or phone or "Customer",
            email=email,
            phone=phone,
        )
        session.add(customer)
        await session.flush()
        counters.customers += 1

        for candidate in key_candidates:
            customers[candidate] = customer

    return customers


async def _import_orders(
    session: AsyncSession,
    rows: Iterable[Mapping[str, Any]],
    counters: ImportCounters,
    customers: dict[str, domain.Customer],
    items: dict[str, domain.Item],
    locations: dict[str, domain.Location],
) -> None:
    for row in rows:
        external_ref = _coerce_str(row.get("external_ref"))
        customer = await _ensure_customer(
            session,
            customers,
            counters,
            name=_coerce_str(row.get("customer_name")),
            email=_coerce_str(row.get("customer_email")),
            phone=_coerce_str(row.get("customer_phone")),
        )

        sale_status = _clean_order_status(row.get("status"))
        sale_date = _coerce_datetime(row.get("order_date")) or utc_now()
        created_at = _coerce_datetime(row.get("created_at")) or sale_date
        subtotal = _coerce_decimal(row.get("subtotal")) or Decimal("0")
        tax = _coerce_decimal(row.get("tax")) or Decimal("0")
        total = _coerce_decimal(row.get("total")) or subtotal + tax
        deposit = _coerce_decimal(row.get("deposit_amt")) or Decimal("0")
        created_by = _coerce_str(row.get("created_by")) or "import.orders"

        sale = domain.Sale(
            customer_id=customer.customer_id if customer else None,
            status=sale_status,
            sale_date=sale_date,
            subtotal=subtotal,
            tax=tax,
            total=total,
            deposit_amt=deposit,
            created_by=created_by,
            source="imported_spreadsheet",
            external_ref=external_ref,
            delivery_requested=False,
            delivery_status=None,
        )
        sale.created_at = created_at
        session.add(sale)
        await session.flush()
        counters.sales += 1

        item_sku = _coerce_str(row.get("item_sku"))
        if item_sku:
            item = items.get(item_sku.lower())
            if item is None:
                counters.warnings.append(
                    f"Skipped order line for unknown product SKU '{item_sku}'"
                )
            else:
                qty = _coerce_decimal(row.get("qty")) or Decimal("1")
                unit_price = _coerce_decimal(row.get("unit_price")) or item.price
                location_name = _coerce_str(row.get("location_name"))
                location = None
                if location_name:
                    location = await _get_or_create_location(
                        session, location_name, locations, counters
                    )
                elif locations:
                    location = next(iter(locations.values()))
                else:
                    location = await _get_or_create_location(
                        session, "Main Warehouse", locations, counters
                    )

                sale_line = domain.SaleLine(
                    sale_id=sale.sale_id,
                    item_id=item.item_id,
                    location_id=location.location_id,
                    qty=qty,
                    unit_price=unit_price,
                    discount=Decimal("0"),
                    tax=Decimal("0"),
                )
                session.add(sale_line)


async def _import_purchase_orders(
    session: AsyncSession,
    rows: Iterable[Mapping[str, Any]],
    counters: ImportCounters,
    vendors: dict[str, domain.Vendor],
    items: dict[str, domain.Item],
) -> None:
    purchase_orders: dict[str, domain.PurchaseOrder] = {}

    for row in rows:
        external_ref = _coerce_str(row.get("external_ref"))
        vendor_name = _coerce_str(row.get("vendor_name")) or "Imported Vendor"
        vendor = await _get_or_create_vendor(session, vendor_name, vendors, counters)

        po_key = (external_ref or vendor_name).lower()
        po = purchase_orders.get(po_key)
        if po is None:
            status_value = _clean_po_status(row.get("status"))
            expected_date = _coerce_datetime(row.get("expected_date"))
            created_at = _coerce_datetime(row.get("created_at"))
            created_by = _coerce_str(row.get("created_by")) or "import.purchase_orders"
            terms = _coerce_str(row.get("terms"))
            notes = _coerce_str(row.get("notes"))

            po = domain.PurchaseOrder(
                vendor_id=vendor.vendor_id,
                status=status_value,
                expected_date=expected_date,
                terms=terms,
                notes=notes,
                created_by=created_by,
                external_ref=external_ref,
            )
            if created_at is not None:
                po.created_at = created_at
            session.add(po)
            await session.flush()
            purchase_orders[po_key] = po
            counters.purchase_orders += 1

        item_sku = _coerce_str(row.get("item_sku"))
        if not item_sku:
            continue
        item = items.get(item_sku.lower())
        if item is None:
            counters.warnings.append(
                f"Skipped purchase order line for unknown product SKU '{item_sku}'"
            )
            continue

        description = _coerce_str(row.get("item_description")) or item.description
        qty = _coerce_decimal(row.get("qty_ordered")) or Decimal("0")
        unit_cost = _coerce_decimal(row.get("unit_cost")) or item.unit_cost

        line = domain.POLine(
            po_id=po.po_id,
            item_id=item.item_id,
            description=description,
            qty_ordered=qty,
            qty_received=Decimal("0"),
            unit_cost=unit_cost,
        )
        session.add(line)


async def _clear_existing_data(session: AsyncSession) -> bool:
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)

    demo_items = await session.scalar(
        select(domain.Item.item_id).where(domain.Item.sku.like("DEMO%"))
    )
    demo_vendor = await session.scalar(
        select(domain.Vendor.vendor_id).where(domain.Vendor.name == "Demo Furnishings")
    )
    cleared_demo = bool(demo_items or demo_vendor)

    models_in_delete_order = [
        domain.IncomingTruckUpdate,
        domain.IncomingTruckLine,
        domain.IncomingTruck,
        domain.ReceivingLine,
        domain.Receiving,
        domain.InventoryTxn,
        domain.POLine,
        domain.PurchaseOrder,
        domain.SaleLine,
        domain.Sale,
        domain.Inventory,
        domain.Barcode,
        domain.Customer,
        domain.Item,
        domain.Location,
        domain.Vendor,
    ]
    for model in models_in_delete_order:
        await session.execute(delete(model))

    await session.flush()

    return cleared_demo


async def _ensure_schema() -> None:
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.create_all)


async def _load_vendor_index(session: AsyncSession) -> dict[str, domain.Vendor]:
    vendors = (await session.scalars(select(domain.Vendor))).all()
    return {vendor.name.lower(): vendor for vendor in vendors}


async def _load_location_index(session: AsyncSession) -> dict[str, domain.Location]:
    locations = (await session.scalars(select(domain.Location))).all()
    return {location.name.lower(): location for location in locations}


async def _load_item_index(
    session: AsyncSession,
) -> tuple[dict[str, domain.Item], set[str]]:
    items = (await session.scalars(select(domain.Item))).all()
    index: dict[str, domain.Item] = {}
    short_codes: set[str] = set()
    for item in items:
        index[item.sku.lower()] = item
        if item.short_code:
            short_codes.add(item.short_code)
    return index, short_codes


async def _load_customer_index(session: AsyncSession) -> dict[str, domain.Customer]:
    customers = (await session.scalars(select(domain.Customer))).all()
    index: dict[str, domain.Customer] = {}
    for customer in customers:
        for key in _customer_lookup_keys(customer.name, customer.email, customer.phone):
            index[key] = customer
    return index


async def _get_or_create_vendor(
    session: AsyncSession,
    vendor_name: str,
    vendor_index: dict[str, domain.Vendor],
    counters: ImportCounters,
) -> domain.Vendor:
    key = vendor_name.lower()
    vendor = vendor_index.get(key)
    if vendor is not None:
        return vendor

    vendor = domain.Vendor(
        name=vendor_name,
        terms=None,
        phone=None,
        email=None,
        address_json=None,
    )
    session.add(vendor)
    await session.flush()
    vendor_index[key] = vendor
    counters.vendors += 1
    return vendor


async def _get_or_create_location(
    session: AsyncSession,
    location_name: str,
    location_index: dict[str, domain.Location],
    counters: ImportCounters,
) -> domain.Location:
    key = location_name.lower()
    location = location_index.get(key)
    if location is not None:
        return location

    location = domain.Location(name=location_name, type="warehouse")
    session.add(location)
    await session.flush()
    location_index[key] = location
    counters.locations += 1
    return location


async def _ensure_customer(
    session: AsyncSession,
    customers: dict[str, domain.Customer],
    counters: ImportCounters,
    *,
    name: str | None,
    email: str | None,
    phone: str | None,
) -> domain.Customer | None:
    for key in _customer_lookup_keys(name, email, phone):
        customer = customers.get(key)
        if customer is not None:
            return customer

    if not any([name, email, phone]):
        return None

    customer = domain.Customer(
        name=name or email or phone or "Customer",
        email=email,
        phone=phone,
    )
    session.add(customer)
    await session.flush()
    counters.customers += 1

    for key in _customer_lookup_keys(customer.name, customer.email, customer.phone):
        customers[key] = customer

    return customer


def extract_datasets(
    data: bytes, filename: str, preferred_entity: str | None = None
) -> dict[str, list[dict[str, Any]]]:
    if not filename.lower().endswith(".xlsx"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Unsupported file type. Upload an XLSX spreadsheet.",
        )
    if load_workbook is None:  # pragma: no cover - optional dependency
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="XLSX support requires the 'openpyxl' package",
        )

    workbook = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)

    for worksheet in workbook.worksheets:
        rows_iter = worksheet.iter_rows(values_only=True)
        headers = None
        normalised_headers: list[str] = []
        for candidate in rows_iter:
            if candidate is None:
                continue
            normalised_candidate = [
                _normalise_header(str(header)) if header is not None else ""
                for header in candidate
            ]
            if not _row_has_values(candidate):
                continue
            if not _row_matches_supported_headers(normalised_candidate):
                continue
            headers = candidate
            normalised_headers = normalised_candidate
            break
        if headers is None:
            continue

        entity_key = _identify_entity(
            worksheet.title, normalised_headers, preferred_entity=preferred_entity
        )
        if entity_key is None:
            continue

        aliases = FIELD_ALIASES.get(entity_key, {})

        for row in rows_iter:
            raw_row: dict[str, Any] = {}
            empty = True
            for index, header in enumerate(normalised_headers):
                if not header:
                    continue
                value = row[index] if index < len(row) else None
                if _has_cell_value(value):
                    empty = False
                raw_row[header] = value
            if empty:
                continue
            grouped[entity_key].append(_prepare_row(entity_key, raw_row))

    return grouped


def _prepare_row(entity: str, raw_row: Mapping[str, Any]) -> dict[str, Any]:
    aliases = FIELD_ALIASES.get(entity, {})
    prepared: dict[str, Any] = {}

    for header, value in raw_row.items():
        field = _resolve_field(aliases, header)
        if field is None:
            continue
        if field in prepared and (value is None or (isinstance(value, str) and not value.strip())):
            continue
        prepared[field] = value

    return prepared


def _resolve_field(aliases: Mapping[str, set[str]], header: str) -> str | None:
    for field, candidates in aliases.items():
        if header in candidates:
            return field
    return None


def _identify_entity(
    title: str, headers: Iterable[str], preferred_entity: str | None = None
) -> str | None:
    normalised_title = _normalise_header(title)
    if normalised_title in SUPPORTED_SHEETS:
        return normalised_title

    scored_entities: dict[str, int] = {}
    for entity, aliases in FIELD_ALIASES.items():
        score = sum(1 for header in headers if _resolve_field(aliases, header))
        if score:
            scored_entities[entity] = score

    if not scored_entities:
        return None

    best_score = max(scored_entities.values())

    if preferred_entity:
        preferred_score = scored_entities.get(preferred_entity)
        if preferred_score and preferred_score == best_score:
            return preferred_entity

    best_entities = [
        entity for entity, score in scored_entities.items() if score == best_score
    ]
    if len(best_entities) == 1:
        return best_entities[0]

    if preferred_entity and preferred_entity in best_entities:
        return preferred_entity

    title_matches = [
        entity
        for entity in best_entities
        if _title_suggests_entity(normalised_title, entity)
    ]
    if len(title_matches) == 1:
        return title_matches[0]

    return None


def _title_suggests_entity(title: str, entity: str) -> bool:
    if not title:
        return False

    if entity in title:
        return True

    singular_entity = entity[:-1] if entity.endswith("s") else entity
    if singular_entity and singular_entity in title:
        return True

    compact_entity = entity.replace("_", "")
    if compact_entity and compact_entity in title:
        return True

    compact_singular = singular_entity.replace("_", "")
    if compact_singular and compact_singular in title:
        return True

    return False


def _row_matches_supported_headers(headers: Iterable[str]) -> bool:
    for header in headers:
        for aliases in FIELD_ALIASES.values():
            if _resolve_field(aliases, header):
                return True
    return False


def _normalise_header(value: str) -> str:
    value = value.strip().lower()
    return re.sub(r"[^a-z0-9]+", "_", value)


def _row_has_values(row: Iterable[Any]) -> bool:
    return any(_has_cell_value(value) for value in row)


def _has_cell_value(value: Any) -> bool:
    if value is None:
        return False
    if isinstance(value, str):
        return bool(value.strip())
    return True


def _coerce_str(value: Any) -> str | None:
    if value is None:
        return None
    if isinstance(value, str):
        value = value.strip()
        return value or None
    if isinstance(value, (int, float, Decimal)):
        return str(value)
    return None


def _coerce_decimal(value: Any) -> Decimal | None:
    if value in (None, ""):
        return None
    if isinstance(value, Decimal):
        return value
    if isinstance(value, (int, float)):
        return Decimal(str(value))
    if isinstance(value, str):
        candidate = value.strip()
        if not candidate:
            return None
        try:
            return Decimal(candidate)
        except InvalidOperation:
            return None
    return None


def _coerce_datetime(value: Any) -> datetime | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        if value.tzinfo is None:
            return value.replace(tzinfo=timezone.utc)
        return value
    if isinstance(value, date):
        return datetime.combine(value, datetime.min.time(), tzinfo=timezone.utc)
    if isinstance(value, (int, float)):
        return datetime.fromtimestamp(float(value), tz=timezone.utc)
    if isinstance(value, str):
        candidate = value.strip()
        if not candidate:
            return None
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%Y-%m-%d %H:%M", "%Y-%m-%dT%H:%M:%S"):
            try:
                parsed = datetime.strptime(candidate, fmt)
                return parsed.replace(tzinfo=timezone.utc)
            except ValueError:
                continue
    return None


def _generate_short_code(sku: str, in_use: set[str]) -> str:
    base = re.sub(r"[^A-Za-z0-9]", "", sku).upper() or "ITEM"
    base = (base + "XXXX")[:4]
    candidate = base[:4]
    suffix = 1
    while candidate in in_use:
        prefix = (base[:2] or "IT").ljust(2, "X")
        candidate = f"{prefix}{suffix:02d}"[-4:]
        suffix += 1
    in_use.add(candidate)
    return candidate


def _customer_lookup_keys(name: str | None, email: str | None, phone: str | None) -> list[str]:
    keys: list[str] = []
    if email:
        keys.append(email.lower())
    if phone:
        keys.append(phone.lower())
    if name:
        keys.append(name.lower())
    return keys


def _clean_order_status(value: Any) -> str:
    candidate = (_coerce_str(value) or "open").lower()
    allowed = {"draft", "open", "fulfilled", "void"}
    if candidate not in allowed:
        return "open"
    return candidate


def _clean_po_status(value: Any) -> str:
    candidate = (_coerce_str(value) or "open").lower()
    allowed = {"draft", "open", "partial", "received", "closed"}
    if candidate not in allowed:
        return "open"
    return candidate
