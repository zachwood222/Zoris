import io
import os

os.environ['DATABASE_URL'] = 'sqlite+aiosqlite:///./test.db'

import pytest
from openpyxl import Workbook
from sqlalchemy import func, select

from .. import sample_data
from ..db import SessionLocal, engine
from ..models.base import Base
from ..models.domain import (
    Customer,
    Inventory,
    Item,
    Location,
    POLine,
    PurchaseOrder,
    Sale,
    SaleLine,
    Vendor,
)
from ..services.importer import NO_IMPORTABLE_ROWS_WARNING, extract_datasets

XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


def _save_workbook(workbook: Workbook) -> io.BytesIO:
    buffer = io.BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return buffer


def _build_products_workbook(
    rows: list[tuple[object, ...]] | None = None,
) -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Products"
    sheet.append(
        [
            "Product",
            "Vendor Model",
            "Description",
            "Descriptions 2",
            "Vend",
            "Type",
            "Cost",
            "Sell Price",
            "Qty On Hand",
        ]
    )
    default_rows: list[tuple[object, ...]] = [
        (
            "SOFA-001",
            "ACME-SOFA-001",
            "Modern Sofa",
            "Gray Fabric",
            "Acme Furniture",
            "Seating",
            450.00,
            899.00,
            4,
        ),
        (
            "LAMP-002",
            "ACME-LAMP-002",
            "Brass Floor Lamp",
            "Matte Finish",
            "Acme Furniture",
            "Lighting",
            60.00,
            129.00,
            6,
        ),
    ]
    for row in rows or default_rows:
        sheet.append(list(row))
    return workbook


def _build_customers_workbook() -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Customers"
    sheet.append(["Name", "Email", "Phone"])
    sheet.append(["Jamie Smith", "jamie@example.com", "555-0100"])
    sheet.append(["Chris Doe", "", "555-0101"])
    return workbook


def _build_orders_workbook() -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Orders"
    sheet.append(
        [
            "Order Number",
            "Customer Email",
            "Customer Name",
            "Item SKU",
            "Qty",
            "Unit Price",
            "Status",
        ]
    )
    sheet.append(
        [
            "ORDER-10",
            "jamie@example.com",
            "Jamie Smith",
            "SOFA-001",
            1,
            899.00,
            "Open",
        ]
    )
    return workbook


def _build_purchase_orders_workbook() -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Purchase Orders"
    sheet.append(
        [
            "PO Number",
            "Vendor Name",
            "Item SKU",
            "Qty Ordered",
            "Unit Cost",
        ]
    )
    sheet.append([
        "PO-50",
        "Acme Furniture",
        "SOFA-001",
        2,
        450.00,
    ])
    return workbook


def _build_vendors_workbook() -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Vendors"
    sheet.append(
        [
            "Vendor Name",
            "Email",
            "Phone",
            "Address",
            "City",
            "State",
            "Postal Code",
            "Country",
            "Terms",
        ]
    )
    sheet.append(
        [
            "Acme Furniture",
            "hello@acme.test",
            "555-0111",
            "123 Market St",
            "Springfield",
            "IL",
            "62701",
            "USA",
            "Net 30",
        ]
    )
    sheet.append(
        [
            "Metro Lighting",
            "contact@metro.test",
            "555-0112",
            "99 Industrial Way",
            "Columbus",
            "OH",
            "43004",
            "USA",
            "Prepaid",
        ]
    )
    return workbook


def _build_blank_products_workbook() -> Workbook:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Products"
    sheet.append(["", "", ""])
    sheet.append([None, None, None])
    sheet.append(["Products", "for", "Import"])
    sheet.append(
        [
            "Product",
            "Vendor Model",
            "Description",
            "Descriptions 2",
            "Vend",
            "Type",
            "Cost",
            "Sell Price",
            "Qty On Hand",
        ]
    )
    sheet.append(["TABLE-01", "MAN-TABLE-01", "Dining Table", "Oak", "Metro", "Dining", 350.0, 499.0, 2])
    return workbook


@pytest.mark.asyncio
async def test_import_products_and_customers(client) -> None:
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.drop_all)

    products_buffer = _save_workbook(_build_products_workbook())
    files = {"file": ("products.xlsx", products_buffer, XLSX_MIME)}
    response = await client.post(
        "/imports/spreadsheet?dataset=products&replaceInventory=true", files=files
    )
    assert response.status_code == 200

    payload = response.json()
    assert payload["counters"]["items"] == 2
    assert payload["counters"]["vendors"] == 1
    assert payload["counters"]["locations"] == 1
    assert payload["counters"]["inventoryRecords"] == 2
    assert payload["clearedInventory"] is True

    customers_buffer = _save_workbook(_build_customers_workbook())
    files = {"file": ("customers.xlsx", customers_buffer, XLSX_MIME)}
    response = await client.post("/imports/spreadsheet?dataset=customers", files=files)
    assert response.status_code == 200

    payload = response.json()
    assert payload["counters"]["customers"] == 2
    assert payload["clearedSampleData"] is False

    async with SessionLocal() as session:
        item_count = await session.scalar(select(func.count(Item.item_id)))
        customer_count = await session.scalar(select(func.count(Customer.customer_id)))
        location_names = set((await session.scalars(select(Location.name))).all())
        inventory_rows = (await session.scalars(select(Inventory))).all()

    assert item_count == 2
    assert customer_count == 2
    assert location_names == {"Main Warehouse"}
    assert {row.qty_on_hand for row in inventory_rows} == {4, 6}

    async with SessionLocal() as session:
        items = (await session.scalars(select(Item))).all()

    descriptions = {item.sku: item.description for item in items}
    assert descriptions["SOFA-001"] == "Modern Sofa - Gray Fabric"
    assert descriptions["LAMP-002"] == "Brass Floor Lamp - Matte Finish"


@pytest.mark.asyncio
async def test_import_products_replaces_inventory_when_requested(client) -> None:
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.drop_all)

    products_buffer = _save_workbook(_build_products_workbook())
    files = {"file": ("products.xlsx", products_buffer, XLSX_MIME)}
    response = await client.post("/imports/spreadsheet?dataset=products", files=files)
    assert response.status_code == 200

    updated_rows = [
        (
            "SOFA-001",
            "ACME-SOFA-001",
            "Modern Sofa",
            "Updated Upholstery",
            "Acme Furniture",
            "Seating",
            475.00,
            929.00,
            8,
        )
    ]
    updated_buffer = _save_workbook(_build_products_workbook(rows=updated_rows))
    files = {"file": ("products-update.xlsx", updated_buffer, XLSX_MIME)}
    response = await client.post(
        "/imports/spreadsheet?dataset=products&replaceInventory=true", files=files
    )
    assert response.status_code == 200
    payload = response.json()
    assert payload["clearedInventory"] is True
    assert payload["counters"]["items"] == 0
    assert payload["counters"]["inventoryRecords"] == 1

    async with SessionLocal() as session:
        inventory_rows = (await session.scalars(select(Inventory))).all()

    assert len(inventory_rows) == 1
    assert float(inventory_rows[0].qty_on_hand) == 8.0


@pytest.mark.asyncio
async def test_import_vendors_only(client) -> None:
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.drop_all)

    vendors_buffer = _save_workbook(_build_vendors_workbook())
    files = {"file": ("vendors.xlsx", vendors_buffer, XLSX_MIME)}
    response = await client.post("/imports/spreadsheet?dataset=vendors", files=files)
    assert response.status_code == 200

    payload = response.json()
    assert payload["counters"]["vendors"] == 2

    update_workbook = Workbook()
    sheet = update_workbook.active
    sheet.title = "Vendors"
    sheet.append(["Vendor Name", "Email", "Phone"])
    sheet.append(["Acme Furniture", "accounts@acme.test", "555-0222"])
    update_buffer = _save_workbook(update_workbook)
    files = {"file": ("vendors-update.xlsx", update_buffer, XLSX_MIME)}
    response = await client.post("/imports/spreadsheet?dataset=vendors", files=files)
    assert response.status_code == 200
    payload = response.json()
    assert payload["counters"]["vendors"] == 0

    async with SessionLocal() as session:
        vendors = (await session.scalars(select(Vendor))).all()

    assert len(vendors) == 2
    vendor_lookup = {vendor.name: vendor for vendor in vendors}
    assert vendor_lookup["Acme Furniture"].email == "accounts@acme.test"
    assert vendor_lookup["Acme Furniture"].phone == "555-0222"


@pytest.mark.asyncio
async def test_import_ignores_leading_blank_rows(client) -> None:
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.drop_all)

    buffer = _save_workbook(_build_blank_products_workbook())
    files = {"file": ("blanks.xlsx", buffer, XLSX_MIME)}
    response = await client.post("/imports/spreadsheet?dataset=products", files=files)
    assert response.status_code == 200

    payload = response.json()
    assert payload["counters"]["items"] == 1


@pytest.mark.asyncio
async def test_import_orders_and_purchase_orders(client) -> None:
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.drop_all)

    products_buffer = _save_workbook(_build_products_workbook())
    files = {"file": ("products.xlsx", products_buffer, XLSX_MIME)}
    await client.post("/imports/spreadsheet?dataset=products", files=files)

    customers_buffer = _save_workbook(_build_customers_workbook())
    files = {"file": ("customers.xlsx", customers_buffer, XLSX_MIME)}
    await client.post("/imports/spreadsheet?dataset=customers", files=files)

    orders_buffer = _save_workbook(_build_orders_workbook())
    files = {"file": ("orders.xlsx", orders_buffer, XLSX_MIME)}
    response = await client.post("/imports/spreadsheet?dataset=orders", files=files)
    assert response.status_code == 200
    assert response.json()["counters"]["sales"] == 1

    po_buffer = _save_workbook(_build_purchase_orders_workbook())
    files = {"file": ("po.xlsx", po_buffer, XLSX_MIME)}
    response = await client.post(
        "/imports/spreadsheet?dataset=purchase_orders", files=files
    )
    assert response.status_code == 200
    assert response.json()["counters"]["purchaseOrders"] == 1

    async with SessionLocal() as session:
        sale = await session.scalar(select(Sale).where(Sale.external_ref == "ORDER-10"))
        po = await session.scalar(select(PurchaseOrder).where(PurchaseOrder.external_ref == "PO-50"))
        sale_lines = []
        po_lines = []
        if sale is not None:
            sale_lines = (
                await session.scalars(select(SaleLine).where(SaleLine.sale_id == sale.sale_id))
            ).all()
        if po is not None:
            po_lines = (
                await session.scalars(select(POLine).where(POLine.po_id == po.po_id))
            ).all()

    assert sale is not None
    assert len(sale_lines) == 1
    assert sale_lines[0].qty == 1
    assert po is not None
    assert len(po_lines) == 1
    assert po_lines[0].qty_ordered == 2


@pytest.mark.asyncio
async def test_import_clears_sample_data(client) -> None:
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.drop_all)

    await sample_data.apply()

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Products"
    sheet.append(
        [
            "Product",
            "Vendor Model",
            "Description",
            "Descriptions 2",
            "Vend",
            "Type",
            "Cost",
            "Sell Price",
            "Qty On Hand",
        ]
    )
    sheet.append([
        "NEW-ITEM",
        "NEW-MODEL",
        "Imported Item",
        "",
        "Acme",
        "Seating",
        120.0,
        199.0,
        0,
    ])

    buffer = _save_workbook(workbook)
    files = {"file": ("fresh.xlsx", buffer, XLSX_MIME)}
    response = await client.post("/imports/spreadsheet?dataset=products", files=files)
    assert response.status_code == 200

    payload = response.json()
    assert payload["clearedSampleData"] is True

    async with SessionLocal() as session:
        demo_item = await session.scalar(select(Item).where(Item.sku.like("DEMO%")))
        imported = await session.scalar(select(Item).where(Item.sku == "NEW-ITEM"))

    assert demo_item is None
    assert imported is not None


@pytest.mark.asyncio
async def test_import_rejects_unsupported_file_types(client) -> None:
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.drop_all)

    files = {"file": ("data.txt", io.BytesIO(b"not-a-spreadsheet"), "text/plain")}

    response = await client.post("/imports/spreadsheet", files=files)
    assert response.status_code == 400
    assert response.json()["detail"].startswith("Unsupported file type")


@pytest.mark.asyncio
async def test_import_returns_warning_when_no_importable_rows(client) -> None:
    async with engine.begin() as conn:
        await conn.run_sync(Base.metadata.drop_all)

    workbook = Workbook()
    notes = workbook.active
    notes.title = "Notes"
    notes.append(["No data here"])

    buffer = _save_workbook(workbook)
    files = {"file": ("notes.xlsx", buffer, XLSX_MIME)}

    response = await client.post("/imports/spreadsheet", files=files)
    assert response.status_code == 200

    payload = response.json()
    assert payload["message"] == NO_IMPORTABLE_ROWS_WARNING
    assert payload["detail"] == NO_IMPORTABLE_ROWS_WARNING
    assert payload["counters"]["warnings"] == [NO_IMPORTABLE_ROWS_WARNING]


def test_extract_datasets_uses_title_to_break_entity_ties() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Vendor List"
    sheet.append(["Name", "Email", "Phone"])
    sheet.append(["Acme Furniture", "sales@acme.test", "555-0100"])

    buffer = _save_workbook(workbook)
    datasets = extract_datasets(buffer.getvalue(), "upload.xlsx")

    assert "vendors" in datasets
    assert len(datasets["vendors"]) == 1
    assert datasets["vendors"][0]["name"] == "Acme Furniture"

